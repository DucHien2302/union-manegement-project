"""
Excel Export Service
Provides functionality to export data to Excel files with formatting
"""

import pandas as pd
from datetime import datetime
from typing import List, Any, Dict, Optional
import os
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from tkinter import filedialog, messagebox
import logging

logger = logging.getLogger(__name__)


class ExcelExportService:
    """Service for exporting data to Excel files with professional formatting"""

    @staticmethod
    def export_members_to_excel(members: List[Any], filename: Optional[str] = None) -> str:
        """
        Export members to Excel file with formatting
        
        Args:
            members: List of member objects
            filename: Optional custom filename
            
        Returns:
            str: Path to the exported file
        """
        try:
            # Prepare data
            data = []
            for member in members:
                # Extract member data
                member_type_str = member.member_type.value if hasattr(member.member_type, 'value') else str(member.member_type)
                status_str = member.status.value if hasattr(member.status, 'value') else str(member.status)
                
                # Format dates
                birth_date = ""
                if hasattr(member, 'date_of_birth') and member.date_of_birth:
                    try:
                        if isinstance(member.date_of_birth, datetime):
                            birth_date = member.date_of_birth.strftime('%d/%m/%Y')
                        else:
                            birth_date = str(member.date_of_birth)
                    except:
                        birth_date = ""
                
                join_date = ""
                if hasattr(member, 'join_date') and member.join_date:
                    try:
                        if isinstance(member.join_date, datetime):
                            join_date = member.join_date.strftime('%d/%m/%Y')
                        else:
                            join_date = str(member.join_date)
                    except:
                        join_date = ""
                
                # Map to Vietnamese display
                type_mapping = {
                    'union_member': '👤 Đoàn viên',
                    'association_member': '👥 Hội viên', 
                    'executive': '👔 Ban chấp hành'
                }
                
                status_mapping = {
                    'active': '✅ Hoạt động',
                    'inactive': '⏸️ Tạm ngưng',
                    'suspended': '❌ Đình chỉ'
                }
                
                data.append({
                    'ID': getattr(member, 'id', ''),
                    'Mã thành viên': getattr(member, 'member_code', ''),
                    'Họ và tên': getattr(member, 'full_name', ''),
                    'Ngày sinh': birth_date,
                    'Giới tính': getattr(member, 'gender', ''),
                    'Số điện thoại': getattr(member, 'phone', ''),
                    'Email': getattr(member, 'email', ''),
                    'Địa chỉ': getattr(member, 'address', ''),
                    'Chức vụ': getattr(member, 'position', ''),
                    'Phòng ban': getattr(member, 'department', ''),
                    'Loại thành viên': type_mapping.get(member_type_str, member_type_str),
                    'Trạng thái': status_mapping.get(status_str, status_str),
                    'Ngày gia nhập': join_date,
                    'Ghi chú': getattr(member, 'notes', '')
                })
            
            # Create DataFrame
            df = pd.DataFrame(data)
            
            # Generate filename if not provided
            if not filename:
                timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
                filename = f"ThanhVien_Export_{timestamp}.xlsx"
            
            # Ask user for save location
            file_path = filedialog.asksaveasfilename(
                defaultextension=".xlsx",
                filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")],
                initialfile=filename,
                title="Lưu file thành viên Excel"
            )
            
            if not file_path:
                return ""
            
            # Create workbook with formatting
            wb = Workbook()
            ws = wb.active
            ws.title = "Thành viên"
            
            # Add title
            ws.merge_cells('A1:N1')
            title_cell = ws['A1']
            title_cell.value = f"DANH SÁCH THÀNH VIÊN - {datetime.now().strftime('%d/%m/%Y')}"
            title_cell.font = Font(name='Arial', size=16, bold=True, color='FFFFFF')
            title_cell.fill = PatternFill(start_color='2E7D32', end_color='2E7D32', fill_type='solid')
            title_cell.alignment = Alignment(horizontal='center', vertical='center')
            
            # Add headers
            headers = list(df.columns)
            for col_num, header in enumerate(headers, 1):
                cell = ws.cell(row=2, column=col_num, value=header)
                cell.font = Font(name='Arial', size=11, bold=True, color='FFFFFF')
                cell.fill = PatternFill(start_color='388E3C', end_color='388E3C', fill_type='solid')
                cell.alignment = Alignment(horizontal='center', vertical='center')
            
            # Add data
            for row_num, row_data in enumerate(df.values, 3):
                for col_num, value in enumerate(row_data, 1):
                    cell = ws.cell(row=row_num, column=col_num, value=value)
                    cell.font = Font(name='Arial', size=10)
                    cell.alignment = Alignment(horizontal='left', vertical='center')
                    
                    # Alternate row colors
                    if row_num % 2 == 0:
                        cell.fill = PatternFill(start_color='F2F2F2', end_color='F2F2F2', fill_type='solid')
                    
                    # Color coding for status
                    if col_num == 12:  # Status column
                        if 'Hoạt động' in str(value):
                            cell.font = Font(name='Arial', size=10, color='2E7D32')
                        elif 'Tạm ngưng' in str(value):
                            cell.font = Font(name='Arial', size=10, color='F57C00')
                        elif 'Đình chỉ' in str(value):
                            cell.font = Font(name='Arial', size=10, color='D32F2F')
            
            # Set column widths
            column_widths = [8, 15, 25, 12, 10, 15, 25, 30, 20, 20, 18, 15, 15, 30]
            for col_num, width in enumerate(column_widths, 1):
                ws.column_dimensions[chr(64 + col_num)].width = width
            
            # Add borders
            thin_border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            
            for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=len(headers)):
                for cell in row:
                    cell.border = thin_border
            
            # Save file
            wb.save(file_path)
            
            logger.info(f"Exported {len(members)} members to {file_path}")
            messagebox.showinfo("Thành công", f"Đã xuất {len(members)} thành viên ra file Excel!\nĐường dẫn: {file_path}")
            
            return file_path
            
        except Exception as e:
            logger.error(f"Error exporting members to Excel: {e}")
            messagebox.showerror("Lỗi", f"Không thể xuất file Excel: {str(e)}")
            return ""

    @staticmethod
    def export_reports_to_excel(reports: List[Any], filename: Optional[str] = None) -> str:
        """
        Export reports to Excel file with formatting
        
        Args:
            reports: List of report objects
            filename: Optional custom filename
            
        Returns:
            str: Path to the exported file
        """
        try:
            # Prepare data
            data = []
            for report in reports:
                # Extract report data
                report_type_str = report.report_type.value if hasattr(report.report_type, 'value') else str(report.report_type)
                status_str = report.status.value if hasattr(report.status, 'value') else str(report.status)
                
                # Format dates
                created_date = ""
                if hasattr(report, 'created_at') and report.created_at:
                    try:
                        if isinstance(report.created_at, datetime):
                            created_date = report.created_at.strftime('%d/%m/%Y')
                        else:
                            created_date = str(report.created_at)
                    except:
                        created_date = ""
                elif hasattr(report, 'created_date') and report.created_date:
                    try:
                        if isinstance(report.created_date, datetime):
                            created_date = report.created_date.strftime('%d/%m/%Y')
                        else:
                            created_date = str(report.created_date)
                    except:
                        created_date = ""
                
                updated_date = ""
                if hasattr(report, 'updated_at') and report.updated_at:
                    try:
                        if isinstance(report.updated_at, datetime):
                            updated_date = report.updated_at.strftime('%d/%m/%Y')
                        else:
                            updated_date = str(report.updated_at)
                    except:
                        updated_date = ""
                
                # Map to Vietnamese display
                type_mapping = {
                    'monthly': '📊 Tháng',
                    'quarterly': '📈 Quý', 
                    'yearly': '📋 Năm',
                    'special': '⭐ Đặc biệt'
                }
                
                status_mapping = {
                    'draft': 'Nháp',
                    'submitted': 'Đã nộp',
                    'approved': 'Đã duyệt',
                    'rejected': 'Từ chối',
                    'in_review': 'Đang xem xét'
                }
                
                # Get creator name - fallback to formatted ID if name not available
                creator_info = ""
                if hasattr(report, 'created_by_name') and report.created_by_name:
                    creator_info = report.created_by_name
                elif hasattr(report, 'created_by') and report.created_by:
                    creator_info = f"User {report.created_by}"
                else:
                    creator_info = "User None"
                
                data.append({
                    'ID': getattr(report, 'id', ''),
                    'Tiêu đề': getattr(report, 'title', ''),
                    'Loại báo cáo': type_mapping.get(report_type_str, report_type_str),
                    'Kỳ báo cáo': getattr(report, 'period', ''),
                    'Trạng thái': status_mapping.get(status_str, status_str),
                    'Người tạo': creator_info,
                    'Ngày tạo': created_date,
                    'Ngày cập nhật': updated_date,
                    'Mô tả': getattr(report, 'description', '')
                })
            
            # Create DataFrame
            df = pd.DataFrame(data)
            
            # Generate filename if not provided
            if not filename:
                timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
                filename = f"BaoCao_Export_{timestamp}.xlsx"
            
            # Ask user for save location
            file_path = filedialog.asksaveasfilename(
                defaultextension=".xlsx",
                filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")],
                initialfile=filename,
                title="Lưu file báo cáo Excel"
            )
            
            if not file_path:
                return ""
            
            # Create workbook with formatting
            wb = Workbook()
            ws = wb.active
            ws.title = "Báo cáo"
            
            # Add title
            ws.merge_cells('A1:I1')
            title_cell = ws['A1']
            title_cell.value = f"DANH SÁCH BÁO CÁO - {datetime.now().strftime('%d/%m/%Y')}"
            title_cell.font = Font(name='Arial', size=16, bold=True, color='FFFFFF')
            title_cell.fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
            title_cell.alignment = Alignment(horizontal='center', vertical='center')
            
            # Add headers
            headers = list(df.columns)
            for col_num, header in enumerate(headers, 1):
                cell = ws.cell(row=2, column=col_num, value=header)
                cell.font = Font(name='Arial', size=11, bold=True, color='FFFFFF')
                cell.fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
                cell.alignment = Alignment(horizontal='center', vertical='center')
            
            # Add data
            for row_num, row_data in enumerate(df.values, 3):
                for col_num, value in enumerate(row_data, 1):
                    cell = ws.cell(row=row_num, column=col_num, value=value)
                    cell.font = Font(name='Arial', size=10)
                    cell.alignment = Alignment(horizontal='left', vertical='center')
                    
                    # Alternate row colors
                    if row_num % 2 == 0:
                        cell.fill = PatternFill(start_color='F2F2F2', end_color='F2F2F2', fill_type='solid')
                    
                    # Color coding for status column (column 5)
                    if col_num == 5:  # Status column
                        if 'Đã duyệt' in str(value):
                            cell.font = Font(name='Arial', size=10, color='2E7D32')
                        elif 'Nháp' in str(value):
                            cell.font = Font(name='Arial', size=10, color='FF9800')
                        elif 'Đã nộp' in str(value):
                            cell.font = Font(name='Arial', size=10, color='2196F3')
                        elif 'Từ chối' in str(value):
                            cell.font = Font(name='Arial', size=10, color='D32F2F')
            
            # Set column widths
            column_widths = [8, 40, 20, 15, 15, 20, 15, 15, 50]
            for col_num, width in enumerate(column_widths, 1):
                ws.column_dimensions[chr(64 + col_num)].width = width
            
            # Add borders
            thin_border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            
            for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=len(headers)):
                for cell in row:
                    cell.border = thin_border
            
            # Save file
            wb.save(file_path)
            
            logger.info(f"Exported {len(reports)} reports to {file_path}")
            messagebox.showinfo("Thành công", f"Đã xuất {len(reports)} báo cáo ra file Excel!\nĐường dẫn: {file_path}")
            
            return file_path
            
        except Exception as e:
            logger.error(f"Error exporting reports to Excel: {e}")
            messagebox.showerror("Lỗi", f"Không thể xuất file Excel: {str(e)}")
            return ""

    @staticmethod
    def export_tasks_to_excel(tasks: List[Any], filename: Optional[str] = None) -> str:
        """
        Export tasks to Excel file with formatting
        
        Args:
            tasks: List of task objects
            filename: Optional custom filename
            
        Returns:
            str: Path to the exported file
        """
        try:
            # Prepare data
            data = []
            for task in tasks:
                # Extract task data
                priority_str = task.priority.value if hasattr(task.priority, 'value') else str(task.priority)
                status_str = task.status.value if hasattr(task.status, 'value') else str(task.status)
                
                # Format dates
                due_date = ""
                if hasattr(task, 'due_date') and task.due_date:
                    try:
                        if isinstance(task.due_date, datetime):
                            due_date = task.due_date.strftime('%d/%m/%Y')
                        else:
                            due_date = str(task.due_date)
                    except:
                        due_date = ""
                
                created_date = ""
                if hasattr(task, 'created_date') and task.created_date:
                    try:
                        if isinstance(task.created_date, datetime):
                            created_date = task.created_date.strftime('%d/%m/%Y')
                        else:
                            created_date = str(task.created_date)
                    except:
                        created_date = ""
                
                # Map to Vietnamese display
                priority_mapping = {
                    'low': '🟢 Thấp',
                    'medium': '🟡 Trung bình',
                    'high': '🟠 Cao',
                    'urgent': '🔴 Khẩn cấp'
                }
                
                status_mapping = {
                    'not_started': '⏸️ Chưa bắt đầu',
                    'in_progress': '⚡ Đang thực hiện',
                    'completed': '✅ Hoàn thành',
                    'cancelled': '❌ Hủy bỏ',
                    'on_hold': '⏸️ Tạm dừng'
                }
                
                # Calculate progress
                progress = f"{getattr(task, 'progress_percentage', 0)}%"
                
                data.append({
                    'ID': getattr(task, 'id', ''),
                    'Tiêu đề': getattr(task, 'title', ''),
                    'Mô tả': getattr(task, 'description', ''),
                    'Độ ưu tiên': priority_mapping.get(priority_str, priority_str),
                    'Trạng thái': status_mapping.get(status_str, status_str),
                    'Người thực hiện': getattr(task, 'assigned_to', ''),
                    'Hạn hoàn thành': due_date,
                    'Tiến độ': progress,
                    'Ngày tạo': created_date
                })
            
            # Create DataFrame
            df = pd.DataFrame(data)
            
            # Generate filename if not provided
            if not filename:
                timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
                filename = f"CongViec_Export_{timestamp}.xlsx"
            
            # Ask user for save location
            file_path = filedialog.asksaveasfilename(
                defaultextension=".xlsx",
                filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")],
                initialfile=filename,
                title="Lưu file công việc Excel"
            )
            
            if not file_path:
                return ""
            
            # Create workbook with formatting
            wb = Workbook()
            ws = wb.active
            ws.title = "Công việc"
            
            # Add title
            ws.merge_cells('A1:I1')
            title_cell = ws['A1']
            title_cell.value = f"DANH SÁCH CÔNG VIỆC - {datetime.now().strftime('%d/%m/%Y')}"
            title_cell.font = Font(name='Arial', size=16, bold=True, color='FFFFFF')
            title_cell.fill = PatternFill(start_color='70AD47', end_color='70AD47', fill_type='solid')
            title_cell.alignment = Alignment(horizontal='center', vertical='center')
            
            # Add headers
            headers = list(df.columns)
            for col_num, header in enumerate(headers, 1):
                cell = ws.cell(row=2, column=col_num, value=header)
                cell.font = Font(name='Arial', size=11, bold=True, color='FFFFFF')
                cell.fill = PatternFill(start_color='548235', end_color='548235', fill_type='solid')
                cell.alignment = Alignment(horizontal='center', vertical='center')
            
            # Add data
            for row_num, row_data in enumerate(df.values, 3):
                for col_num, value in enumerate(row_data, 1):
                    cell = ws.cell(row=row_num, column=col_num, value=value)
                    cell.font = Font(name='Arial', size=10)
                    cell.alignment = Alignment(horizontal='left', vertical='center')
                    
                    # Alternate row colors
                    if row_num % 2 == 0:
                        cell.fill = PatternFill(start_color='F2F2F2', end_color='F2F2F2', fill_type='solid')
                    
                    # Color coding for status
                    if col_num == 5:  # Status column
                        if 'Hoàn thành' in str(value):
                            cell.font = Font(name='Arial', size=10, color='2E7D32')
                        elif 'Đang thực hiện' in str(value):
                            cell.font = Font(name='Arial', size=10, color='1565C0')
                        elif 'Hủy bỏ' in str(value):
                            cell.font = Font(name='Arial', size=10, color='D32F2F')
            
            # Set column widths
            column_widths = [8, 35, 40, 15, 18, 20, 15, 12, 15]
            for col_num, width in enumerate(column_widths, 1):
                ws.column_dimensions[chr(64 + col_num)].width = width
            
            # Add borders
            thin_border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            
            for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=len(headers)):
                for cell in row:
                    cell.border = thin_border
            
            # Save file
            wb.save(file_path)
            
            logger.info(f"Exported {len(tasks)} tasks to {file_path}")
            messagebox.showinfo("Thành công", f"Đã xuất {len(tasks)} công việc ra file Excel!\nĐường dẫn: {file_path}")
            
            return file_path
            
        except Exception as e:
            logger.error(f"Error exporting tasks to Excel: {e}")
            messagebox.showerror("Lỗi", f"Không thể xuất file Excel: {str(e)}")
            return ""

    @staticmethod
    def export_filtered_data(data: List[Dict], data_type: str, filename: Optional[str] = None) -> str:
        """
        Export filtered data to Excel
        
        Args:
            data: List of dictionaries containing the data
            data_type: Type of data ('reports' or 'tasks')
            filename: Optional custom filename
            
        Returns:
            str: Path to the exported file
        """
        try:
            if data_type == 'reports':
                # Convert dict data back to report-like objects for processing
                # This is a simplified approach - in real implementation,
                # you might want to create proper data classes
                return ExcelExportService.export_reports_to_excel(data, filename)
            elif data_type == 'tasks':
                return ExcelExportService.export_tasks_to_excel(data, filename)
            else:
                raise ValueError(f"Unsupported data type: {data_type}")
                
        except Exception as e:
            logger.error(f"Error exporting filtered data: {e}")
            messagebox.showerror("Lỗi", f"Không thể xuất dữ liệu đã lọc: {str(e)}")
            return ""